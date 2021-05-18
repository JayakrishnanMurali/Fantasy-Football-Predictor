import asyncio
import json
import pandas as pd
import openpyxl
import http.client
import aiohttp
from openpyxl.styles import (PatternFill, colors)
from understat import Understat

#Change Value Here
team1="Manchester United"
team2="Fulham"
league="epl"

teamList = [team1,team2]


async def main():
    async with aiohttp.ClientSession() as session:
        understat = Understat(session)

         # ***********************GENERATE***FIXTURE ID***************************** #

        wb_obj = openpyxl.load_workbook('Fixtures.xlsx')
        sheet_obj = wb_obj.active

        for i in range (1,sheet_obj.max_row+1):
            for j in range (1,11):
                cell_obj = sheet_obj.cell(row=i, column=j)
                if (cell_obj.value == team1):
                    fixture_id = str(sheet_obj.cell(row=i, column=2).value)
                    break


         # ***********************LEAGUE***TABLE***************************** #

        league_table = await understat.get_league_table("epl", "2021")

        lt = pd.DataFrame(league_table)
        
        
        # print (lt)
        lt.to_excel('LeagueTable.xlsx', sheet_name='Sheet1')

        wb_obj = openpyxl.load_workbook('LeagueTable.xlsx')
        sheet_obj = wb_obj.active

        for i in range (1,sheet_obj.max_row+1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if (cell_obj.value == team1 or cell_obj.value == team2):
                cell_obj.fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
                
        wb_obj.save('LeagueTable.xlsx')

        # ***********************TEAM******STATS*********************************** #
        for i in range(len(teamList)):
            results = await understat.get_team_results(
                teamList[i],
                2021
                
            )
            teamstat = json.dumps(results)
            js_data = json.loads(teamstat)

            ts = pd.DataFrame()
            dict = {}
            
            for data in js_data:
                home_side = data['h']['title']
                away_side = data['a']['title']
                home_score = data['goals']['h']
                away_score = data['goals']['a']
                score = str(home_score)+'-'+str(away_score)

                dict = {"home_side":home_side,"away_side":away_side,"score":score}
                ts = ts.append(dict,ignore_index=True)

         
            ts = ts[['home_side', 'away_side', 'score']]  
    
            
            ts.to_excel(teamList[i]+'.xlsx', sheet_name='Sheet1')



        # ***********************PLAYER******STATS*********************************** #

        stat = await understat.get_league_players(league, 2021, team_title=team1)
        stat2 = await understat.get_league_players(league, 2021, team_title=team2)
        
        # Create DataFrame
        df = pd.DataFrame()
        
        # Attach two sheets
        df = df.append(stat2,ignore_index=True)
        df = df.append(stat,ignore_index=True)
        
        # Convert to Number
       
        df['games'] = pd.to_numeric(df['games'])
        df['goals'] = pd.to_numeric(df['goals'])
        df['xG'] = pd.to_numeric(df['xG'])
        df['assists'] = pd.to_numeric(df['assists'])
        df['xA'] = pd.to_numeric(df['xA'])
        df['shots'] = pd.to_numeric(df['shots'])
        df['key_passes'] = pd.to_numeric(df['key_passes'])
        df['xGChain'] = pd.to_numeric(df['xGChain'])
        df['xGBuildup'] = pd.to_numeric(df['xGBuildup'])
        df['time'] = pd.to_numeric(df['time'])


        # Drop Columns not needed
        df = df.drop(columns =['id','yellow_cards','red_cards','npg','npxG'])

        # Sort columns 
        df = df.sort_values(['assists', 'goals', 'key_passes', 'shots'], ascending=[False,False,False,False])


        # print (df)
        df.to_excel('Players.xlsx', sheet_name='Sheet1')

        
        # ************************LINE******UP********************************* #
        conn = http.client.HTTPSConnection("v3.football.api-sports.io")

        headers = {
            'x-rapidapi-host': "v3.football.api-sports.io",
            'x-rapidapi-key': "*************************************"
            }

        conn.request("GET", "/fixtures/lineups?fixture="+fixture_id, headers=headers)

        res = conn.getresponse()
        lineUp = res.read()

        lineUp = str(lineUp, 'UTF-8')

        js_lineup = json.loads(lineUp)

        df = pd.DataFrame()

        inmatch = []
        for players in js_lineup['response']:
            for p in players['startXI']:
                name = str(p['player']['name'])
                inmatch.append(name)
                print(name)

        print (inmatch)
        print ("Success!")


if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())
